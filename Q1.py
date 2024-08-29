######### Important:Make sure the excel file is not running in the background during the execution of the script#######

import openpyxl
from selenium import webdriver
import time
from selenium.webdriver.common.by import By
from datetime import date

today = date.today()

today = today.strftime("%A")
# today="Sunday"
# This function is for returning max and min result
def minmax(search):
    driver = webdriver.Chrome()

    # Maximize the window
    driver.maximize_window()

    # Navigate to the URL
    driver.get("https://www.google.com/")

    # Identify the Google search text box and enter the value
    # search = "javatpoint"
    search_box = driver.find_element(By.NAME, "q")
    search_box.send_keys(search)


    time.sleep(4)
    # Capture the suggestions
    suggestions = driver.find_elements(By.XPATH, "//ul[@role='listbox']//li//div[@class='wM6W7d']/span")

    # Print all suggestions
    print("Search Suggestions:")
    trimed_reslts = [suggestion.text.strip() for suggestion in suggestions if suggestion.text.strip()]
    # for trimed in trimed_reslts:
    #     print(trimed)
    print(min(trimed_reslts, key=len))
    driver.close()
    return min(trimed_reslts, key=len),max(trimed_reslts, key=len)
# Load the workbook
workbook = openpyxl.load_workbook(r'E:\dragons\4BeatsQ1.xlsx')

# Select the active sheet
sheet = workbook.active
# Matching Today with the tab of excel file
for sheet_name in workbook.sheetnames:
    if sheet_name == today:
        mysheet=sheet_name
        break
print(mysheet)
sheet = workbook[mysheet]

start_cell=sheet.cell(row=1, column=4)
print(start_cell)
if start_cell.value is None :
    print("started at 2 4")
    start=3
else:
    print("started at 1 4")
    start=2
count = start
print(sheet.max_row)
for row in sheet.iter_rows(min_row=start, max_row=sheet.max_row - 1, min_col=2, values_only=False):
    print(row)


    # cell = row[1]
    # cell_value = cell.value

    # print(count)
    # print(f'{count}  : {row[1]}')
    finish_cell = sheet.cell(row=count, column=2)
    # print(finish_cell)
    if finish_cell.value is None:
        # print("started at 2 4")
        break
    result = minmax(row[1].value)
    # print(result)
    sheet.cell(row=count,column=4,value=result[1]) # Writing into longest cell
    sheet.cell(row=count, column=5, value=result[0]) #writting into shrtest cell
    count += 1
workbook.save(r'E:\dragons\4BeatsQ1.xlsx')